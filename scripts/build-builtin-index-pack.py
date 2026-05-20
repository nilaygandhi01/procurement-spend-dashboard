#!/usr/bin/env python3
"""Scan data/inputs/indexes/*.xlsx and emit the JSON pack that the dashboard's
inline IDP_BUILTIN_INDEX_PACK consumes.

We accept two xlsx layouts (both observed in this repo):

 (A) FRED-style: README + Monthly tabs. Monthly tab has header row
     ("observation_date", "<SERIES_CODE>") followed by N monthly rows of
     date + numeric value. README tab last row is the canonical title.

 (B) Sparse-sample: single tab "Sheet" with title in A1, "Value" in B1, then
     date + value rows (possibly only 3-10 observations). Used for the
     synthetic placeholder workbooks shipped alongside the real FRED ones.

For each file we compute the yearly average across whatever monthly /
sub-yearly observations are present. The dashboard rebases against 2024
at render time, so a file that only has 2024-2026 data still works as long
as 2024 is represented.

The pack we emit matches the existing IDP_BUILTIN_INDEX_PACK schema:

    {
      "<CODE>": {
        "displayName": "...",        # human-readable label for the checkbox
        "rawByYear":   {YYYY: value, ...},
        "blsSeriesId": "WPU10",      # informational, also used as the key
        "sourceFile":  "WPU10.xlsx"  # informational, where the data came from
      },
      ...
    }

Run:
    py scripts/build-builtin-index-pack.py            # prints JSON to stdout
    py scripts/build-builtin-index-pack.py --write     # writes generated.json
"""
from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import datetime

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl not installed. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


INDEXES_DIR = os.path.join("data", "inputs", "indexes")


def _is_date(v):
    return isinstance(v, datetime)


def _series_from_fred_sheet(ws):
    """FRED 'Monthly' tab: header row + monthly observations. Return
    {YYYY: avg_value} and the series code from the header."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None, None
    header = rows[0]
    series_code = str(header[1]).strip() if len(header) > 1 and header[1] else None
    yearly: dict[int, list[float]] = {}
    for r in rows[1:]:
        if not r or len(r) < 2 or not _is_date(r[0]):
            continue
        try:
            val = float(r[1])
        except (TypeError, ValueError):
            continue
        yearly.setdefault(r[0].year, []).append(val)
    avg = {y: round(sum(vs) / len(vs), 6) for y, vs in yearly.items() if vs}
    return avg, series_code


def _series_from_sparse_sheet(ws):
    """Single-tab sparse layout: title in A1, "Value" in B1, then date+value
    rows. Same yearly-average reduction."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None, None
    title = str(rows[0][0]).strip() if rows[0] and rows[0][0] else None
    yearly: dict[int, list[float]] = {}
    for r in rows[1:]:
        if not r or len(r) < 2 or not _is_date(r[0]):
            continue
        try:
            val = float(r[1])
        except (TypeError, ValueError):
            continue
        yearly.setdefault(r[0].year, []).append(val)
    avg = {y: round(sum(vs) / len(vs), 6) for y, vs in yearly.items() if vs}
    return avg, title


def _fred_readme_title(wb):
    """Last row of README tab carries the series description in BLS pattern:
    (CODE, "Producer Price Index … Index <base>, …", "Data Updated: ...")."""
    if "README" not in wb.sheetnames:
        return None
    ws = wb["README"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None
    for r in reversed(rows):
        if r and len(r) >= 2 and r[1]:
            return str(r[1]).strip()
    return None


def process_workbook(path: str) -> dict | None:
    """Return one IDP_BUILTIN_INDEX_PACK entry, or None on failure."""
    name = os.path.basename(path)
    code = os.path.splitext(name)[0]
    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        print(f"[warn] {name}: load_workbook failed: {e}", file=sys.stderr)
        return None
    raw_by_year = None
    series_code = None
    display = None
    if "Monthly" in wb.sheetnames:
        raw_by_year, series_code = _series_from_fred_sheet(wb["Monthly"])
        display = _fred_readme_title(wb) or series_code or code
        # Shorten the display name for the checkbox label.
        if display and len(display) > 90:
            display = display.split(",")[0].strip()
    else:
        ws = wb.active
        raw_by_year, display = _series_from_sparse_sheet(ws)
        if not display:
            display = code
        series_code = code
    if not raw_by_year:
        print(f"[warn] {name}: no usable yearly data extracted", file=sys.stderr)
        return None
    return {
        "code": series_code or code,
        "displayName": display,
        "rawByYear": raw_by_year,
        "blsSeriesId": series_code or code,
        "sourceFile": name,
        "pointCount": sum(1 for _ in raw_by_year),
    }


def build_pack(directory: str) -> dict:
    pack: dict[str, dict] = {}
    if not os.path.isdir(directory):
        print(f"[error] not a directory: {directory}", file=sys.stderr)
        return pack
    for fname in sorted(os.listdir(directory)):
        if not fname.lower().endswith(".xlsx"):
            continue
        entry = process_workbook(os.path.join(directory, fname))
        if entry is None:
            continue
        key = entry["code"]
        if key in pack:
            print(f"[warn] {fname}: duplicate series code {key}; keeping first", file=sys.stderr)
            continue
        pack[key] = entry
    return pack


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dir", default=INDEXES_DIR, help="folder of .xlsx files")
    ap.add_argument("--write", action="store_true", help="write data/inputs/index-data/generated-index-pack.json")
    args = ap.parse_args()
    pack = build_pack(args.dir)
    text = json.dumps(pack, indent=2, sort_keys=True)
    if args.write:
        out = os.path.join("data", "inputs", "index-data", "generated-index-pack.json")
        with open(out, "w", encoding="utf-8") as fh:
            fh.write(text)
        print(f"Wrote {out} with {len(pack)} entries:", file=sys.stderr)
        for k, v in pack.items():
            print(f"  {k:20} {len(v['rawByYear']):4d}y  {v['displayName'][:60]}", file=sys.stderr)
    else:
        print(text)


if __name__ == "__main__":
    main()
